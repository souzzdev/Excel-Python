import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import re

def formatar_cpf(cpf):
    """
    Formata um CPF para o padrão ###.###.###-##
    
    Args:
        cpf (str): CPF a ser formatado.
    
    Returns:
        str: CPF formatado.
    """
    cpf = re.sub(r'\D', '', cpf)  # Remove caracteres não numéricos
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"

def formatar_cnpj(cnpj):
    """
    Formata um CNPJ para o padrão ##.###.###/####-##
    
    Args:
        cnpj (str): CNPJ a ser formatado.
    
    Returns:
        str: CNPJ formatado.
    """
    cnpj = re.sub(r'\D', '', cnpj)  # Remove caracteres não numéricos
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def validar_email(email):
    """
    Valida se um e-mail está em um formato correto.
    
    Args:
        email (str): E-mail a ser validado.
    
    Returns:
        bool: True se o e-mail é válido, False caso contrário.
    """
    regex = r'^[\w\.-]+@[a-zA-Z\d\.-]+\.[a-zA-Z]{2,}$'
    return re.match(regex, email) is not None

def formatar_telefone(telefone):
    """
    Formata um número de telefone para o padrão (##) #####-####
    
    Args:
        telefone (str): Número de telefone a ser formatado.
    
    Returns:
        str: Telefone formatado.
    """
    telefone = re.sub(r'\D', '', telefone)  # Remove caracteres não numéricos
    if len(telefone) == 11:
        return f"({telefone[:2]}) {telefone[2:7]}-{telefone[7:]}"
    elif len(telefone) == 10:
        return f"({telefone[:2]}) {telefone[2:6]}-{telefone[6:]}"
    else:
        return telefone  # Retorna o número sem formatação se não corresponder aos padrões conhecidos

def criar_tabela_gerenciamento():
    """
    Cria uma tabela de gerenciamento formatada e amigável em um arquivo Excel (.xlsx).
    """
    # Cria um novo workbook e seleciona a planilha ativa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gerenciamento"

    # Define o cabeçalho da tabela
    cabecalho = []
    while True:
        col_name = input("Adicione o nome da célula (ou deixe em branco para terminar): ").strip()
        if not col_name:
            break
        cabecalho.append(col_name)

    ws.append(cabecalho)

    # Loop para entrada de dados
    while True:
        linha_dados = []
        for col_name in cabecalho:
            valor = input(f"Digite o {col_name} (ou para várias linhas, separe por ';'): ").strip()
            # Separa o valor de entrada para múltiplas células
            linha_dados.append(valor.split(';'))
        if all(not valor for linha in linha_dados for valor in linha):
            break
        # Achata os valores múltiplos em uma lista única
        linha_dados = [item for sublist in linha_dados for item in sublist]
        ws.append(linha_dados)

    # Define o intervalo da tabela
    table_ref = f"A1:{chr(64 + len(cabecalho))}{ws.max_row}"

    # Cria a tabela com filtro automático
    tab = Table(displayName="TabelaGerenciamento", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    ws.add_table(tab)

    # Formatação personalizada do cabeçalho
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde Escuro
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Formatação das células restantes com fundo Verde Claro e texto Branco
    rest_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde Claro
    rest_font = Font(color="000000")  # Texto Preto
    alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cabecalho)):
        for cell in row:
            cell.fill = rest_fill
            cell.font = rest_font
            cell.alignment = alignment

            # Formatação específica baseada no nome da coluna
            if cabecalho[cell.column - 1].lower() == "cpf":
                cell.value = formatar_cpf(cell.value)
            elif cabecalho[cell.column - 1].lower() == "cnpj":
                cell.value = formatar_cnpj(cell.value)
            elif cabecalho[cell.column - 1].lower() == "telefone":
                cell.value = formatar_telefone(cell.value)
            elif cabecalho[cell.column - 1].lower() == "email":
                if not validar_email(cell.value):
                    cell.font = Font(color="FF0000")  # Vermelho para e-mails inválidos

    # Ajusta a largura das colunas com base no valor mais longo
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                if isinstance(cell.value, list):
                    for value in cell.value:
                        if len(str(value)) > max_length:
                            max_length = len(str(value))
                else:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Adiciona bordas
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Protege a planilha se o usuário optar por colocar senha
    proteger_planilha = input("Deseja proteger a planilha com senha? (sim/não): ").strip().lower()
    if proteger_planilha == "sim":
        senha = input("Digite a senha desejada para proteger a planilha: ").strip()
        ws.protection.sheet = True
        ws.protection.password = senha

    # Salva o workbook como arquivo .xlsx
    nome_arquivo = input("Digite o nome do arquivo Excel a ser criado (sem a extensão .xlsx): ").strip()
    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"
    wb.save(nome_arquivo)

    print(f"Tabela de gerenciamento criada com sucesso no arquivo '{nome_arquivo}'!")

# Chama a função para criar a tabela de gerenciamento
criar_tabela_gerenciamento()